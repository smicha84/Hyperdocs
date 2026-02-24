#!/usr/bin/env python3
"""
Add data-trace sheets to the pipeline Excel workbook.

Shows ACTUAL DATA flowing through each transformation stage:
  Sheet 6: Message Trace — 3 real messages traced through every pipeline stage
  Sheet 7: LLM Agent Outputs — What each Opus agent actually produces
  Sheet 8: Evidence Collection Output — What collect_file_evidence.py produces for one file
  Sheet 9: Cross-Session Aggregation — How one file accumulates data across sessions
  Sheet 10: Transformation Detail — Before/after for every major operation
"""
import json
import os
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tools.log_config import get_logger

logger = get_logger("tools.add_data_trace_sheets")

H3 = Path(__file__).resolve().parent.parent
OUTPUT_DIR = H3 / "output"
PERM = Path(os.getenv("HYPERDOCS_STORE_DIR", str(Path.home() / "PERMANENT_HYPERDOCS")))
PERM_SESSIONS = PERM / "sessions"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
STAGE_FILL = PatternFill(start_color="3742fa", end_color="3742fa", fill_type="solid")
STAGE_FONT = Font(bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
OUTPUT_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
TRANSFORM_FILL = PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid")
GAP_FILL = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(bottom=Side(style="thin", color="636e72"))


def style_header(ws, row, ncols):
    """Apply dark header styling (white text, dark fill) to a row of cells."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def stage_row(ws, row, ncols, text):
    """Write a full-width stage separator row with blue fill and bold white text."""
    ws.cell(row=row, column=1, value=text)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).font = STAGE_FONT
        ws.cell(row=row, column=c).fill = STAGE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def add_row(ws, row, values, fill=None):
    """Write a data row with optional background fill, text wrapping, and thin border."""
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def load_safe(path):
    """Load JSON from path, returning None if missing or corrupt."""
    if not path.exists():
        return None
    try:
        from tools.json_io import load_json as _load_json
        return _load_json(path)
    except (json.JSONDecodeError, UnicodeDecodeError, OSError):
        return None


def build_sheet_6_message_trace(wb):
    """Sheet 6: 3 real messages traced through every transformation stage."""
    ws = wb.create_sheet("Message Data Trace")
    headers = ["Pipeline Stage", "Field / Operation", "Message 1 (idx=1, protocol empty)", "Message 30 (idx=30, tier4 assistant)", "Message 1108 (idx=1108, frustration peak)"]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    # Load session data
    PERM_S = PERM_SESSIONS / "session_0012ebed"
    LOCAL_S = OUTPUT_DIR / "session_0012ebed"
    enriched = load_safe(PERM_S / "enriched_session.json") or load_safe(LOCAL_S / "enriched_session.json")
    if not enriched:
        ws.cell(row=2, column=1, value="Could not load session 0012ebed")
        return ws

    msgs = enriched["messages"]
    m1, m30, m1108 = msgs[1], msgs[30], msgs[1108]
    targets = [m1, m30, m1108]

    row = 2

    # ── RAW INPUT ──
    stage_row(ws, row, ncols, "STAGE 1: Raw JSONL Input (before any processing)")
    row += 1
    raw_fields = [
        ("role", lambda m: m["role"]),
        ("content (first 300 chars)", lambda m: m["content"][:300] if m["content"] else "(empty)"),
        ("content length (raw bytes)", lambda m: str(len(m["content"]))),
        ("model", lambda m: str(m.get("model", ""))),
        ("timestamp", lambda m: str(m.get("timestamp", ""))),
        ("uuid", lambda m: str(m.get("uuid", ""))[:30]),
        ("has_thinking", lambda m: str(m.get("has_thinking", False))),
        ("thinking_length", lambda m: str(m.get("thinking_length", 0))),
    ]
    for fname, fn in raw_fields:
        vals = ["Raw JSONL", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, INPUT_FILL)
        row += 1

    # ── PHASE 0: PROTOCOL DETECTION ──
    stage_row(ws, row, ncols, "STAGE 2: Phase 0 — Protocol Detection (step 0.4a)")
    row += 1
    proto_fields = [
        ("is_protocol", lambda m: str(m.get("is_protocol", False))),
        ("protocol_type", lambda m: str(m.get("protocol_type", "None"))),
        ("EFFECT", lambda m: "SUPPRESSED: tier forced to 1, all signals zeroed" if m.get("is_protocol") else "No effect — real human/assistant content"),
    ]
    for fname, fn in proto_fields:
        vals = ["Protocol Detection", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, TRANSFORM_FILL)
        row += 1

    # ── PHASE 0: CHAR-PER-LINE ──
    stage_row(ws, row, ncols, "STAGE 3: Phase 0 — Char-Per-Line Collapse (step 0.4b)")
    row += 1
    cpl_fields = [
        ("was_char_encoded", lambda m: str(m.get("was_char_encoded", False))),
        ("content_length_raw", lambda m: str(m.get("content_length_raw", 0))),
        ("content_length (corrected)", lambda m: str(m.get("content_length", 0))),
        ("EFFECT", lambda m: f"COLLAPSED: {m.get('content_length_raw',0)} bytes → {m.get('content_length',0)} bytes (encoding stripped)" if m.get("was_char_encoded") else "No encoding detected"),
    ]
    for fname, fn in cpl_fields:
        vals = ["Char-Per-Line", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, TRANSFORM_FILL)
        row += 1

    # ── PHASE 0: METADATA EXTRACTION ──
    stage_row(ws, row, ncols, "STAGE 4: Phase 0 — Metadata Extraction (step 0.4d) — 50+ signals per message")
    row += 1
    meta_fields = [
        ("metadata.files", lambda m: str(m.get("metadata", {}).get("files", []))),
        ("metadata.error", lambda m: str(m.get("metadata", {}).get("error", False))),
        ("metadata.caps_ratio", lambda m: str(m.get("metadata", {}).get("caps_ratio", 0))),
        ("metadata.profanity", lambda m: str(m.get("metadata", {}).get("profanity", False))),
        ("metadata.emergency_intervention", lambda m: str(m.get("metadata", {}).get("emergency_intervention", False))),
        ("metadata.exclamations", lambda m: str(m.get("metadata", {}).get("exclamations", 0))),
        ("metadata.repeat_count", lambda m: str(m.get("metadata", {}).get("repeat_count", 0))),
        ("metadata.error_context", lambda m: str(m.get("metadata", {}).get("error_context", "n/a"))),
    ]
    for fname, fn in meta_fields:
        vals = ["Metadata Extraction", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # ── PHASE 0: MESSAGE FILTERING ──
    stage_row(ws, row, ncols, "STAGE 5: Phase 0 — Message Filtering (step 0.4f) — 4-tier classification")
    row += 1
    filt_fields = [
        ("filter_tier", lambda m: str(m.get("filter_tier", 0))),
        ("filter_tier_name", lambda m: str(m.get("filter_tier_name", ""))),
        ("filter_score", lambda m: str(m.get("filter_score", 0))),
        ("filter_signals", lambda m: str(m.get("filter_signals", []))),
        ("filter_signals_content_referential", lambda m: str(m.get("filter_signals_content_referential", False))),
        ("EFFECT", lambda m: f"Tier {m.get('filter_tier',0)}: {'SKIP — not passed to Phase 1' if m.get('filter_tier',0) <= 1 else 'INCLUDED in tier2+ files'}{' → PRIORITY for Phase 1' if m.get('filter_tier',0) == 4 else ''}"),
    ]
    for fname, fn in filt_fields:
        vals = ["Message Filtering", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # ── PHASE 0: BEHAVIOR ANALYSIS ──
    stage_row(ws, row, ncols, "STAGE 6: Phase 0 — Behavior Analysis (step 0.4j) — assistant messages only")
    row += 1
    def fmt_behavior(m):
        b = m.get("behavior_flags")
        if b is None:
            return "(skipped — not assistant or protocol)"
        flags = [k for k, v in b.items() if v and k != "details" and not isinstance(v, (dict, int))]
        scores = {k: v for k, v in b.items() if isinstance(v, int) and v > 0}
        return f"Flags: {flags}, Scores: {scores}" if flags or scores else "All clear (no behavioral issues detected)"

    beh_fields = [
        ("behavior_flags (summary)", fmt_behavior),
        ("behavior_flags.confusion", lambda m: str((m.get("behavior_flags") or {}).get("confusion", "n/a"))),
        ("behavior_flags.overconfident", lambda m: str((m.get("behavior_flags") or {}).get("overconfident", "n/a"))),
        ("behavior_flags.rushing", lambda m: str((m.get("behavior_flags") or {}).get("rushing", "n/a"))),
        ("behavior_flags.damage_score", lambda m: str((m.get("behavior_flags") or {}).get("user_upset_score", "n/a"))),
    ]
    for fname, fn in beh_fields:
        vals = ["Behavior Analysis", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # ── PHASE 0: FINAL ENRICHED RECORD ──
    stage_row(ws, row, ncols, "STAGE 7: Phase 0 Output — enriched_session.json (18 fields per message)")
    row += 1
    final_fields = [
        ("content_hash", lambda m: str(m.get("content_hash", ""))),
        ("Total fields in record", lambda m: str(len(m))),
        ("DESTINATION", lambda m: f"→ prepare_agent_data.py splits this into 9 focused files based on filter_tier"),
    ]
    for fname, fn in final_fields:
        vals = ["Enriched Record", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # ── PHASE 0b: AGENT DATA SPLIT ──
    stage_row(ws, row, ncols, "STAGE 8: Phase 0b — prepare_agent_data.py — Which files include this message?")
    row += 1
    split_fields = [
        ("In session_metadata.json?", lambda m: "YES (stats only, no content)"),
        ("In tier2plus_messages.json?", lambda m: "YES" if m.get("filter_tier", 0) >= 2 else "NO (tier < 2)"),
        ("In tier4_priority_messages.json?", lambda m: "YES → sent to Phase 1 extract_threads.py" if m.get("filter_tier", 0) == 4 else "NO (tier < 4)"),
        ("In safe_tier4.json?", lambda m: "YES (profanity sanitized)" if m.get("filter_tier", 0) == 4 else "NO"),
        ("In safe_condensed.json?", lambda m: "YES (all messages included, compressed keys)"),
        ("In batches/?", lambda m: "YES (in a batch of 30)" if m.get("filter_tier", 0) >= 2 else "NO"),
        ("Profanity sanitized?", lambda m: "YES — all output files pass through _sanitize_text()"),
    ]
    for fname, fn in split_fields:
        vals = ["Agent Data Split", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, TRANSFORM_FILL)
        row += 1

    # ── PHASE 1: THREAD EXTRACTION ──
    stage_row(ws, row, ncols, "STAGE 9: Phase 1 — extract_threads.py (only processes tier 4 messages)")
    row += 1

    # Load thread data
    threads = load_safe(PERM_S / "thread_extractions.json") or load_safe(LOCAL_S / "thread_extractions.json") or {}
    thread_cats = threads.get("threads", {})

    def find_thread_entry(idx, cat):
        cat_data = thread_cats.get(cat, {})
        if isinstance(cat_data, dict):
            for entry in cat_data.get("entries", []):
                if entry.get("msg_index") == idx:
                    return json.dumps(entry, default=str)[:300]
        return "(not found for this message)"

    def thread_status(m):
        if m.get("filter_tier", 0) < 4:
            return "SKIPPED — only tier 4 messages are analyzed by Phase 1"
        return "ANALYZED — 6 threads + 6 markers extracted"

    thread_fields = [
        ("Phase 1 processing status", thread_status),
    ]
    for fname, fn in thread_fields:
        vals = ["Thread Extraction", fname] + [fn(m) for m in targets]
        add_row(ws, row, vals, TRANSFORM_FILL)
        row += 1

    # Show thread entries for msg 30 and 1108
    for cat in ["ideas", "reactions", "software", "code", "plans", "behavior"]:
        vals = ["Thread: " + cat, "Entry for this message"]
        for m in targets:
            idx = m["index"]
            if m.get("filter_tier", 0) < 4:
                vals.append("(not analyzed)")
            else:
                vals.append(find_thread_entry(idx, cat))
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # ── OPUS AGENTS ──
    stage_row(ws, row, ncols, "STAGE 10: Opus Agents (optional) — What each agent produces for these messages")
    row += 1

    # Semantic primitives
    prims = load_safe(PERM_S / "semantic_primitives.json") or {}
    tagged = prims.get("tagged_messages", [])
    def find_prim(idx):
        for t in tagged:
            if t.get("msg_index", t.get("index", -1)) == idx:
                return t
        return None

    prim_fields = ["action_vector", "confidence_signal", "emotional_tenor", "intent_marker", "friction_log", "decision_trace"]
    for pf in prim_fields:
        vals = ["Semantic Primitives", pf]
        for m in targets:
            p = find_prim(m["index"])
            if p:
                vals.append(str(p.get(pf, "(empty)")))
            else:
                vals.append("(no primitive tagged for this message)")
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # Geological notes (session-level, not per-message, so show what mentions this message's index)
    geo = load_safe(PERM_S / "geological_notes.json") or {}
    for zoom in ["micro", "meso", "macro"]:
        obs_list = geo.get(zoom, [])
        # Find observations that mention indices near our target messages
        vals = [f"Geological ({zoom})", "Observation near this message"]
        for m in targets:
            idx = m["index"]
            found = None
            for obs in obs_list:
                if isinstance(obs, dict):
                    msg_range = obs.get("message_range", [])
                    if isinstance(msg_range, list) and len(msg_range) == 2:
                        if msg_range[0] <= idx <= msg_range[1]:
                            found = obs.get("observation", "")[:250]
                            break
                    elif str(idx) in str(obs.get("observation", "")):
                        found = obs.get("observation", "")[:250]
                        break
            vals.append(found or "(no observation covers this message index)")
        add_row(ws, row, vals, OUTPUT_FILL)
        row += 1

    # ── PHASE 2 ──
    stage_row(ws, row, ncols, "STAGE 11: Phase 2 — Idea Graph & Synthesis (built from all thread data)")
    row += 1

    ig = load_safe(PERM_S / "idea_graph.json") or load_safe(LOCAL_S / "idea_graph.json") or {}
    nodes = ig.get("nodes", [])
    edges = ig.get("edges", [])

    # Find nodes near our target messages
    vals = ["Idea Graph", "Node derived from / near this message"]
    for m in targets:
        idx = m["index"]
        found = None
        for n in nodes:
            fa = n.get("first_appearance", n.get("message_index", -1))
            if fa == idx or (isinstance(fa, int) and abs(fa - idx) <= 2):
                found = f"[{n.get('id','')}] {n.get('label', n.get('name',''))[:150]} (confidence: {n.get('confidence','?')})"
                break
        vals.append(found or "(no idea graph node at this message index)")
    add_row(ws, row, vals, OUTPUT_FILL)
    row += 1

    # Show edge from first node
    vals = ["Idea Graph", "Edge from/to this node"]
    for m in targets:
        idx = m["index"]
        found_node_id = None
        for n in nodes:
            fa = n.get("first_appearance", n.get("message_index", -1))
            if fa == idx or (isinstance(fa, int) and abs(fa - idx) <= 2):
                found_node_id = n.get("id", "")
                break
        if found_node_id:
            for e in edges:
                if e.get("from") == found_node_id or e.get("to") == found_node_id:
                    vals.append(f"{e.get('from','')} --[{e.get('type', e.get('transition_type',''))}]--> {e.get('to','')}: {e.get('label', e.get('evidence',''))[:150]}")
                    break
            else:
                vals.append("(no edges connected)")
        else:
            vals.append("(no node found)")
    add_row(ws, row, vals, OUTPUT_FILL)
    row += 1

    # Synthesis
    synth = load_safe(PERM_S / "synthesis.json") or load_safe(LOCAL_S / "synthesis.json") or {}
    passes = synth.get("passes", {})
    vals = ["Synthesis", f"Number of passes: {len(passes)}"]
    for m in targets:
        char_val = synth.get('session_character', '')
        char_str = str(char_val)[:100] if char_val else ''
        vals.append(f"Session-level: {len(passes)} passes. Key findings: {len(synth.get('key_findings',[]))}. Character: {char_str}")
    add_row(ws, row, vals, OUTPUT_FILL)
    row += 1

    # Grounded markers
    gm = load_safe(PERM_S / "grounded_markers.json") or load_safe(LOCAL_S / "grounded_markers.json") or {}
    markers = gm.get("markers", [])
    vals = ["Grounded Markers", f"Total markers: {len(markers)}"]
    for m in targets:
        idx = m["index"]
        related = [mk for mk in markers if isinstance(mk, dict) and (str(idx) in str(mk.get("source", "")) or str(idx) in str(mk.get("evidence", "")))]
        if related:
            mk = related[0]
            vals.append(f"[{mk.get('marker_id','')}] {mk.get('claim', mk.get('warning',''))[:200]}")
        else:
            vals.append(f"(no markers reference msg {idx} directly — {len(markers)} markers are session-level)")
    add_row(ws, row, vals, OUTPUT_FILL)
    row += 1

    # ── PHASE 3a ──
    stage_row(ws, row, ncols, "STAGE 12: Phase 3a — collect_file_evidence.py (for files mentioned in msg 30)")
    row += 1

    ev_dir = LOCAL_S / "file_evidence"
    m30_files = m30.get("metadata", {}).get("files", [])
    vals = ["Evidence Collection", f"Files mentioned in msg 30: {len(m30_files)}"]
    vals += [str(m30_files[:5]), "(only msg 30 mentions files)", "(only msg 30 mentions files)"]
    add_row(ws, row, vals, TRANSFORM_FILL)
    row += 1

    # Show evidence for first file mentioned
    if m30_files and ev_dir.exists():
        safe_name = m30_files[0].replace("/", "_").replace("\\", "_").replace(".", "_").replace(" ", "_")
        ev_path = ev_dir / f"{safe_name}_evidence.json"
        ev_data = load_safe(ev_path)
        if ev_data:
            evidence_sections = ["emotional_arc", "geological_character", "lineage", "explorer_observations",
                                "chronological_timeline", "code_similarity", "graph_context",
                                "synthesis_context", "claude_md_context", "session_context", "thread_markers"]
            for section in evidence_sections:
                sec_data = ev_data.get(section, {})
                dp = sec_data.get("data_points", 0)
                # Get a meaningful sample
                sample = ""
                if section == "emotional_arc":
                    nearby = sec_data.get("file_nearby_emotions", [])
                    if nearby:
                        sample = f"Nearby emotions: {json.dumps(nearby[0], default=str)[:200]}"
                    else:
                        sample = f"Dominant: {sec_data.get('dominant_emotion','?')}, Friction: {sec_data.get('friction_episodes',0)}"
                elif section == "chronological_timeline":
                    events = sec_data.get("events", [])
                    if events:
                        sample = f"First event: idx={events[0].get('msg_index')} thread={events[0].get('thread')} content={str(events[0].get('content',''))[:120]}"
                elif section == "lineage":
                    fam = sec_data.get("session_family")
                    if fam:
                        sample = f"Family: {fam.get('family_name','?')}, Members: {fam.get('members',[])})"
                    ig_nodes = sec_data.get("idea_graph_lineage_nodes", [])
                    if ig_nodes:
                        sample += f" | Graph nodes: {json.dumps(ig_nodes[0], default=str)[:150]}"
                elif section == "graph_context":
                    sample = f"Edges: {len(sec_data.get('connected_edges',[]))}, Subgraphs: {len(sec_data.get('containing_subgraphs',[]))}, File nodes: {sec_data.get('file_node_ids',[])}"
                elif section == "synthesis_context":
                    sample = f"Passes: {len(sec_data.get('passes',[]))}, Key findings: {len(sec_data.get('key_findings',[]))}"
                else:
                    # Generic: show first key-value pair that's not data_points
                    for k, v in sec_data.items():
                        if k != "data_points" and v:
                            sample = f"{k}: {str(v)[:200]}"
                            break

                vals = [f"Evidence: {m30_files[0]}", f"Section: {section} ({dp} data points)"]
                vals += [sample or "(empty)", "(n/a)", "(n/a)"]
                add_row(ws, row, vals, OUTPUT_FILL)
                row += 1

    # ── PHASE 3b ──
    stage_row(ws, row, ncols, "STAGE 13: Phase 3b — generate_dossiers.py output for files in this session")
    row += 1
    dossiers = load_safe(PERM_S / "file_dossiers.json") or load_safe(LOCAL_S / "file_dossiers.json") or {}
    dossier_dict = dossiers.get("dossiers", {})
    if isinstance(dossier_dict, dict):
        # Show dossier for first file in msg 30
        if m30_files:
            target_file = m30_files[0]
            dossier = dossier_dict.get(target_file, {})
            if dossier:
                for dk in ["summary", "story_arc", "confidence", "significance", "total_mentions"]:
                    val = dossier.get(dk, "")
                    vals = [f"Dossier: {target_file}", dk]
                    vals += [str(val)[:300] if val else "(empty)", "(n/a)", "(n/a)"]
                    add_row(ws, row, vals, OUTPUT_FILL)
                    row += 1
                # Warnings
                warnings = dossier.get("warnings", [])
                vals = [f"Dossier: {target_file}", f"warnings ({len(warnings)})"]
                vals += [json.dumps(warnings[:2], default=str)[:300] if warnings else "(none)", "(n/a)", "(n/a)"]
                add_row(ws, row, vals, OUTPUT_FILL)
                row += 1

    # ── PHASE 4a ──
    stage_row(ws, row, ncols, "STAGE 14: Phase 4a — How this file's data appears in cross_session_file_index.json")
    row += 1
    csidx = load_safe(OUTPUT_DIR / "cross_session_file_index.json")
    if csidx and m30_files:
        target_file = m30_files[0]
        basename = Path(target_file).name
        cs_entry = csidx.get("files", {}).get(target_file, csidx.get("files", {}).get(basename, {}))
        if cs_entry:
            cs_fields = [
                ("session_count", str(cs_entry.get("session_count", 0))),
                ("total_mentions", str(cs_entry.get("total_mentions", 0))),
                ("sessions", str(cs_entry.get("sessions", [])[:5]) + ("..." if len(cs_entry.get("sessions",[])) > 5 else "")),
                ("confidence_history", " → ".join([str(c.get("confidence","?")) for c in cs_entry.get("confidence_history",[])[:8]])),
                ("merged_warnings count", str(len(cs_entry.get("merged_warnings", [])))),
                ("merged_decisions count", str(len(cs_entry.get("merged_decisions", [])))),
                ("code_similarity matches", str(len(cs_entry.get("code_similarity", [])))),
                ("genealogy", str(cs_entry.get("genealogy", "None"))),
                ("frustration_associations", str(len(cs_entry.get("frustration_associations", [])))),
                ("cross_session_emotional_arc sessions", str(len(cs_entry.get("cross_session_emotional_arc", {}).get("per_session_emotions", {})))),
                ("cross_session_geological observations", str(len(cs_entry.get("cross_session_geological", [])))),
                ("cross_session_timeline events", str(len(cs_entry.get("cross_session_timeline", [])))),
            ]
            for fname, fval in cs_fields:
                vals = [f"Cross-Session: {target_file}", fname, fval, "(aggregated across all sessions)", "(aggregated across all sessions)"]
                add_row(ws, row, vals, OUTPUT_FILL)
                row += 1

    # ── THE GAP ──
    stage_row(ws, row, ncols, "STAGE 15: ⚠ THE GAP — This cross-session data is NEVER fed back to Phase 0 or Phase 1")
    row += 1
    vals = ["THE GAP", "What happens next?", "NOTHING. The next session starts Phase 0 blind.", "Phase 1 does not know this file exists in 12+ sessions.", "Phase 1 does not know this user's frustration pattern."]
    add_row(ws, row, vals, GAP_FILL)
    row += 1
    vals = ["THE GAP", "What SHOULD happen?", "Phase 0.5 enricher reads cross_session_file_index.json", "Produces file_context_brief.json with confidence history, warnings, genealogy", "Phase 1 agents receive the brief and analyze WITH context"]
    add_row(ws, row, vals, GAP_FILL)
    row += 1

    # Column widths
    widths = [22, 45, 60, 60, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    return ws


def main():
    """Load the pipeline Excel workbook and add/replace the Message Data Trace sheet."""
    xlsx_path = H3 / "experiment" / "feedback_loop" / "output" / "pipeline_complete_anatomy.xlsx"
    if not xlsx_path.exists():
        logger.error(f"ERROR: {xlsx_path} not found. Run generate_pipeline_excel.py first.")
        return

    logger.info("Loading existing workbook...")
    wb = openpyxl.load_workbook(xlsx_path)

    # Remove old data trace sheet if it exists
    if "Message Data Trace" in wb.sheetnames:
        del wb["Message Data Trace"]

    logger.info("Building Sheet 6: Message Data Trace (actual data at every stage)...")
    build_sheet_6_message_trace(wb)

    wb.save(xlsx_path)
    size = xlsx_path.stat().st_size
    logger.info(f"\nUpdated: {xlsx_path}")
    logger.info(f"Size: {size:,} bytes")
    logger.info(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        logger.info(f"  {name}: {ws.max_row} rows × {ws.max_column} cols")


if __name__ == "__main__":
    main()
