#!/usr/bin/env python3
"""
Generate an exhaustive Excel workbook documenting every detail of the Hyperdocs pipeline.

Reads actual session data and catalogs:
  - Every file in every session, with sizes and field counts
  - Every field in every JSON file, with types, depths, and sample values
  - The complete processing chain: what creates each file, what reads it
  - Per-session statistics across all 50+ processed sessions
  - The cross-session index field-by-field breakdown

Output: pipeline_complete_anatomy.xlsx
"""
import json
import os
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict, OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────
H3 = Path(__file__).resolve().parent.parent
OUTPUT_DIR = H3 / "output"
PERM = Path(os.getenv("HYPERDOCS_STORE_DIR", str(Path.home() / "PERMANENT_HYPERDOCS")))
PERM_SESSIONS = PERM / "sessions"
INDEXES_DIR = PERM / "indexes"

# ── Styles ─────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
PHASE_FILL = PatternFill(start_color="3742fa", end_color="3742fa", fill_type="solid")
PHASE_FONT = Font(bold=True, color="FFFFFF", size=11)
FREE_FILL = PatternFill(start_color="26de81", end_color="26de81", fill_type="solid")
PAID_FILL = PatternFill(start_color="fc5c65", end_color="fc5c65", fill_type="solid")
GAP_FILL = PatternFill(start_color="fc5c65", end_color="fc5c65", fill_type="solid")
GAP_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="57606f", end_color="57606f", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
FIELD_TYPE_COLORS = {
    "str": "dfe6e9", "int": "ffeaa7", "float": "ffeaa7",
    "list": "81ecec", "dict": "a29bfe", "bool": "fab1a0",
    "null": "b2bec3",
}
THIN_BORDER = Border(
    bottom=Side(style="thin", color="636e72"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_phase_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = PHASE_FONT
        cell.fill = PHASE_FILL


def auto_width(ws, min_width=10, max_width=60):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, min(len(line) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def flatten_json_fields(data, prefix="", depth=0, max_depth=4):
    """Recursively flatten a JSON structure into (path, type, sample, depth) tuples."""
    fields = []
    if depth > max_depth:
        return fields

    if isinstance(data, dict):
        for key, val in data.items():
            path = f"{prefix}.{key}" if prefix else key
            val_type = type(val).__name__
            if val_type == "NoneType":
                val_type = "null"

            # Sample value
            if isinstance(val, (str, int, float, bool)):
                sample = str(val)[:200]
            elif isinstance(val, list):
                sample = f"[{len(val)} items]"
                if val and isinstance(val[0], (str, int, float)):
                    sample += f" e.g. {str(val[0])[:80]}"
            elif isinstance(val, dict):
                sample = f"{{{len(val)} keys}}"
                keys_preview = ", ".join(list(val.keys())[:5])
                sample += f" keys: {keys_preview}"
            elif val is None:
                sample = "null"
            else:
                sample = str(val)[:100]

            fields.append((path, val_type, sample, depth))

            # Recurse into dicts and the first item of lists
            if isinstance(val, dict):
                fields.extend(flatten_json_fields(val, path, depth + 1, max_depth))
            elif isinstance(val, list) and val:
                if isinstance(val[0], dict):
                    fields.extend(flatten_json_fields(val[0], f"{path}[0]", depth + 1, max_depth))

    return fields


def load_json_safe(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, UnicodeDecodeError, OSError):
        return None


# ── SHEET 1: Pipeline Steps (every micro-operation) ───────────

PIPELINE_STEPS = [
    # (phase, step_id, script, operation, reads, writes, api, detail)
    ("Phase 0", "0.1", "deterministic_prep.py", "Load session JSONL",
     "{session_id}.jsonl", "", "None ($0)",
     "ClaudeSessionReader.load_session_file() — parses each JSONL line into a ClaudeMessage object with: role, content, timestamp, uuid, model, thinking, tool_calls. Handles Claude Code streaming format where messages are split across multiple JSON lines."),

    ("Phase 0", "0.2", "deterministic_prep.py", "Initialize 3 extractors",
     "", "", "None ($0)",
     "MetadataExtractor() — 50+ signal types per message. MessageFilter(verbose=False) — 4-tier classification (skip/basic/standard/priority). ClaudeBehaviorAnalyzer() — context damage, rushing, overconfidence detection."),

    ("Phase 0", "0.3", "deterministic_prep.py", "Detect subagent session",
     "", "", "None ($0)",
     "detect_subagent_session() — 3 strategies: (1) session ID contains '_agent-' prefix, (2) JSONL filename contains '_agent-', (3) first 5 messages contain 'Hello memory agent' or '<observed_from_primary'. Tags as memory_agent, observer_agent, or task_agent."),

    ("Phase 0", "0.4a", "deterministic_prep.py", "PER MESSAGE: Protocol detection (raw)",
     "", "", "None ($0)",
     "detect_protocol_message() — checks for: empty/whitespace wrappers, 8 XML protocol tag patterns (system-reminder, command-name, command-message, command-args, command-stdout, command-stderr, local-command-caveat, task-notification), /clear continuation boilerplate (3 marker strings), skill injection (.claude/skills/ content), subagent relay markers ('Hello memory agent', 'PROGRESS SUMMARY CHECKPOINT', '<observed_from_primary'). Returns {is_protocol: bool, protocol_type: string}."),

    ("Phase 0", "0.4b", "deterministic_prep.py", "PER MESSAGE: Char-per-line collapse",
     "", "", "None ($0)",
     "collapse_char_per_line() — detects content where each character is on its own line (e.g., 'I\\nm\\np\\nl\\ne\\nm\\ne\\nn\\nt'). If >70% of lines are single characters AND >6 lines total: joins all lines into one string. Returns (collapsed_content, was_encoded). Fixes content_length, caps_ratio, file detection, and profanity detection that would otherwise fail on encoded content."),

    ("Phase 0", "0.4c", "deterministic_prep.py", "PER MESSAGE: Protocol re-check (collapsed)",
     "", "", "None ($0)",
     "Re-runs detect_protocol_message() on collapsed content. Catches /clear continuation markers that were char-per-line encoded and invisible in raw form. Only runs if was_char_encoded=true AND first check was is_protocol=false."),

    ("Phase 0", "0.4d", "deterministic_prep.py", "PER MESSAGE: Metadata extraction (50+ signals)",
     "", "", "None ($0)",
     "claude_to_geological() adapter converts ClaudeMessage → GeologicalMessage. MetadataExtractor.extract_message_metadata(geo_msg, idx) extracts: files (mentioned filenames via regex), files_create (new files), files_edit (modified files), files_open (opened files), error (bool), error_type, caps_ratio (float 0-1), profanity (bool), exclamations (count), emergency_intervention (bool), emergency_reason (string), repeat_count (int), code_blocks (count), tool_calls (list), model (string), content_length (int)."),

    ("Phase 0", "0.4e", "deterministic_prep.py", "PER MESSAGE: Tool failure detection",
     "", "", "None ($0)",
     "For assistant messages with model='<synthetic>' and 'error' in content: marks is_synthetic_error=true, increments session-level tool_failure_count. Synthetic messages are tool result wrappers inserted by Claude Code."),

    ("Phase 0", "0.4f", "deterministic_prep.py", "PER MESSAGE: Message filtering (4-tier classification)",
     "", "", "None ($0)",
     "MessageFilter.classify(analysis_content) — scores content against signal patterns: failure (error keywords), frustration (emotional keywords), breakthrough (success keywords), pivot (direction change), architecture (design discussion), plan (task/checklist). Produces: tier (1=skip, 2=basic, 3=standard, 4=priority), tier_name, score (int), signals (list of 'signal:count' strings)."),

    ("Phase 0", "0.4g", "deterministic_prep.py", "PER MESSAGE: Protocol signal suppression",
     "", "", "None ($0)",
     "If is_protocol=true: force tier=1 (skip) regardless of content richness. Suppress error=false, profanity=false, caps_ratio=0, emergency_intervention=false, emergency_reason=''. Reason: continuation summaries contain quoted content from prior sessions — signals from that text belong to those sessions, not this one. Tags error_context='protocol_recap' if error was true."),

    ("Phase 0", "0.4h", "deterministic_prep.py", "PER MESSAGE: Content-referential detection",
     "", "", "None ($0)",
     "detect_content_referential_signals() — 3 strategies: (1) analytical indicators on assistant messages >500 chars with 2+ analysis keywords (problem, failure mode, gate, enforcer, etc.), (2) signal density anomaly: >20 failure or >10 frustration signals on >1000 char messages, (3) moderate failure+architecture signals on >500 char messages. Also: positive tone + failure signals on assistant messages >300 chars. Returns bool: is this message DISCUSSING errors or EXPERIENCING errors?"),

    ("Phase 0", "0.4i", "deterministic_prep.py", "PER MESSAGE: Error context tagging",
     "", "", "None ($0)",
     "If error=true AND (is_content_referential=true OR content >500 chars) AND model != '<synthetic>': set error_context='mentioned_not_encountered' and error=false. Separates errors the session TALKS about (in analysis reports, code reviews) from errors the session EXPERIENCES (actual runtime failures)."),

    ("Phase 0", "0.4j", "deterministic_prep.py", "PER MESSAGE: Behavior analysis (assistant only)",
     "", "", "None ($0)",
     "ClaudeBehaviorAnalyzer.analyze_message(msg, prev_5_msgs) — detects: overconfident (bool), rushing (bool), ignores_context (bool), confusion (bool), damage_score (0-5). Uses rolling window of last 5-10 assistant messages for temporal context. Skipped entirely for protocol messages (empty wrappers get spurious confusion/damage flags)."),

    ("Phase 0", "0.4k", "deterministic_prep.py", "PER MESSAGE: Build enriched record",
     "", "", "None ($0)",
     "Assembles ONE JSON object per message with 18 fields: index (int), role (str), content (str, FULL — never truncated), content_length (int, corrected for char-per-line), content_length_raw (int, original), content_hash (str, sha256 first 16 chars), timestamp (ISO string), uuid (str), model (str), has_thinking (bool), thinking_length (int), metadata (dict, 50+ signals), filter_tier (int 1-4), filter_tier_name (str), filter_score (int), filter_signals (list of strings), filter_signals_content_referential (bool), behavior_flags (dict or null), is_protocol (bool), protocol_type (str or null), was_char_encoded (bool), llm_behavior (null — populated by optional LLM passes)."),

    ("Phase 0", "0.4l", "deterministic_prep.py", "PER MESSAGE: Accumulate session stats",
     "", "", "None ($0)",
     "Updates running accumulators: tier_distribution (4 counters), frustration_peaks (list of {index, caps_ratio, profanity, content_preview} — only user messages, not protocol, with caps>0.3 or profanity=true), emergency_interventions (list of {index, reason, content_preview} — not protocol), file_mention_counts (dict of filename → count), error_count (int)."),

    ("Phase 0", "0.5", "deterministic_prep.py", "False positive file removal",
     "", "", "None ($0)",
     "Two strategies: (1) blocklist of 17 generic filenames (file.py, test.py, config.py, setup.py, main.py, app.py, utils.py, helper.py, module.py, script.py, run.py, index.js, index.html, style.css, styles.css, file.txt, data.json, output.json). (2) Substring elimination: if a short filename (base ≤12 chars) is a substring of a longer detected filename, remove the short one. Cleans both session_stats.file_mention_counts and per-message metadata.files arrays."),

    ("Phase 0", "0.6", "deterministic_prep.py", "Write enriched_session.json",
     "", "enriched_session.json", "None ($0)",
     "One JSON file containing: session_id (str), source_file (str), generated_at (ISO str), generator ('deterministic_prep.py (Phase 0)'), session_stats (dict with 20+ fields: total_messages, user_messages, assistant_messages, human_messages, protocol_messages, tier_distribution, frustration_peaks, file_mention_counts, error_count, tool_failure_count, emergency_interventions, total_input_tokens, total_output_tokens, total_thinking_chars, is_subagent, agent_id, agent_type, char_per_line_messages, top_files, false_positive_files_removed), messages (array of enriched records). Typical size: 2-8 MB."),

    ("Phase 0b", "0b.1", "prepare_agent_data.py", "Load enriched session",
     "enriched_session_v2.json OR enriched_session.json", "", "None ($0)",
     "Prefers v2 (has LLM pass behavioral tags) if it exists. Falls back to v1 (base Python-only). Reads the full messages array and session_stats."),

    ("Phase 0b", "0b.2", "prepare_agent_data.py", "Extract session metadata",
     "", "session_metadata.json", "None ($0)",
     "Everything from enriched_session.json EXCEPT the messages array. Contains: session_id, source_file, generated_at, generator, session_stats (full). Typically 5-50 KB."),

    ("Phase 0b", "0b.3", "prepare_agent_data.py", "Filter tier 2+ messages",
     "", "tier2plus_messages.json", "None ($0)",
     "All messages where filter_tier >= 2. Full content preserved. Format: {count: N, messages: [...]}. Typically 1-5 MB."),

    ("Phase 0b", "0b.4", "prepare_agent_data.py", "Filter tier 4 priority messages",
     "", "tier4_priority_messages.json", "None ($0)",
     "Only messages where filter_tier == 4 (highest signal). Full content preserved. This is the PRIMARY INPUT for Phase 1 (extract_threads.py). Format: {count: N, messages: [...]}. Typically 0.2-2 MB."),

    ("Phase 0b", "0b.5", "prepare_agent_data.py", "Build conversation condensed",
     "", "conversation_condensed.json", "None ($0)",
     "ALL messages with compressed keys to save space: i (index), r (role), c (content, FULL), cl (content_length), t (filter_tier), ts (timestamp), meta (files, error, caps, profanity, emergency), signals (filter_signals), behavior (behavior_flags), is_protocol, protocol_type, was_char_encoded, content_ref, llm_behavior. Typically 1-5 MB."),

    ("Phase 0b", "0b.6", "prepare_agent_data.py", "Filter user messages tier 2+",
     "", "user_messages_tier2plus.json", "None ($0)",
     "User messages only, tier 2+. For frustration and reaction analysis. Format: {count: N, messages: [...]}. Typically 0.1-1 MB."),

    ("Phase 0b", "0b.7", "prepare_agent_data.py", "Build emergency context windows",
     "", "emergency_contexts.json", "None ($0)",
     "For each emergency_intervention index: a window of ±5 messages centered on the emergency. Format: {count: N, windows: [{emergency_index: int, context_messages: [...]}]}. Typically 1-100 KB."),

    ("Phase 0b", "0b.8", "prepare_agent_data.py", "Split into batch files",
     "", "batches/batch_001.json ... batch_NNN.json", "None ($0)",
     "Tier 2+ messages split into batches of 30. Each batch: {batch_number: int, start_index: int, end_index: int, count: int, messages: [...]}. Typically 5-20 batches per session, ~100 KB each."),

    ("Phase 0b", "0b.9", "prepare_agent_data.py", "Profanity sanitization (pass 1 — disk files)",
     "All *.json in session dir (recursive)", "All *.json (in-place modification)", "None ($0)",
     "Scans ALL JSON files written to disk in the session directory. Replaces 13 profanity words with [expletive]/[Expletive]/[EXPLETIVE]. Also handles keystroke-encoded profanity (f\\nu\\nc\\nk, s\\nh\\ni\\nt, etc.). Ensures all files sent to LLM APIs comply with content policies."),

    ("Phase 0b", "0b.10", "prepare_agent_data.py", "Build safe files (profanity-sanitized, full content)",
     "", "safe_tier4.json + safe_condensed.json", "None ($0)",
     "safe_tier4.json: tier 4 messages with profanity replaced, char-per-line collapsed in content previews. 18 fields per message including llm_behavior. safe_condensed.json: ALL messages, same treatment. These are the PRIMARY FILES agents should read."),

    ("Phase 0b", "0b.11", "prepare_agent_data.py", "Opus-filtered files (conditional)",
     "opus_classifications.json (if exists)", "safe_opus_priority.json + opus_priority_messages.json", "None ($0)",
     "Only runs if opus_classifications.json exists from a prior opus_classifier.py run. Builds Opus-classified priority messages that may differ from Python tier-4 classification. Reports delta: +N new messages, -N dropped. Agents should prefer these when available."),

    ("Schema Norm", "N.1", "schema_normalizer.py", "Normalize thread_extractions.json",
     "thread_extractions.json", "thread_extractions.json (in-place, backup in backups/)", "None ($0)",
     "5 strategies: (1) threads as dict with category sub-keys → extract entries per category. (2) threads as list → group by thread/category field. (3) extractions as list (dominant: 182 sessions) → group by thread/category. (4) extractions as dict → convert to category entries. (5) thread_N_ top-level keys → wrap in category format. Canonical output: {threads: {category: {description: str, entries: list}}, _extra: {everything else}, _normalization_log: [steps taken]}. Backup original to backups/thread_extractions.json. Atomic write via tempfile."),

    ("Schema Norm", "N.2", "schema_normalizer.py", "Normalize semantic_primitives.json",
     "semantic_primitives.json", "semantic_primitives.json (in-place)", "None ($0)",
     "5 strategies: (1) tagged_messages list (already canonical). (2) primitives list (direct or segment-grouped). (3) segments list → flatten. (4) primitives_by_segment dict/list. (5) message_primitives/per_exchange_primitives/tagged_segments/significant_messages. Canonical: {tagged_messages: [], distributions: {emotional_tenor, action_vector, confidence_signal, intent_marker}, summary_statistics: {}, _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.3", "schema_normalizer.py", "Normalize geological_notes.json",
     "geological_notes.json", "geological_notes.json (in-place)", "None ($0)",
     "14 key name variants: micro, meso, macro, observations, strata, layers, geological_layers, zoom_levels, exploration_notes, analysis_levels, vertical_structures, stratigraphy, stratigraphic_column, geological_features, deep_time_observations. Also collects cross_cutting_features, fault_lines, fossils. Canonical: {micro: [], meso: [], macro: [], observations: [], geological_metaphor: str, _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.4", "schema_normalizer.py", "Normalize explorer_notes.json",
     "explorer_notes.json", "explorer_notes.json (in-place)", "None ($0)",
     "12 primary key names + 11 section key names merged into observations. Primary: observations, exploration_notes, explorer_observations, exploration_findings, free_notes, notes, notable_observations, surprising_observations, overlooked_patterns, explorer_findings, explorations, exploration_summary. Sections: abandoned_ideas, patterns, warnings, anomalies, unanswered_questions, open_questions, what_matters_most, emotional_dynamics, cross_session_connections, cross_session_links, ideas_detected. Canonical: {observations: [], explorer_summary: str, _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.5", "schema_normalizer.py", "Normalize idea_graph.json",
     "idea_graph.json", "idea_graph.json (in-place)", "None ($0)",
     "3 strategies: (1) nodes/edges at top level. (2) nested in 'graph' dict. (3) 'idea_nodes' key. Edges from 'edges' or 'transitions'. Canonical: {nodes: [], edges: [], metadata: {}, subgraphs: [] (if present), _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.6", "schema_normalizer.py", "Normalize synthesis.json",
     "synthesis.json", "synthesis.json (in-place)", "None ($0)",
     "5 strategies: (1) passes as dict (dominant: 191 sessions). (2) passes as list. (3) pass_N_ top-level keys. (4) six_pass_synthesis nested. (5) one_sentence/one_line_summary style → consolidate. Canonical: {passes: {}, key_findings: [], session_character: str, cross_session_links: [], _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.7", "schema_normalizer.py", "Normalize grounded_markers.json",
     "grounded_markers.json", "grounded_markers.json (in-place)", "None ($0)",
     "6 strategies: (1) markers list (direct). (2) markers empty → recover from warnings/recommendations/patterns. (3) markers empty + metrics dict → convert to summary marker. (4) no markers key → check session_level_markers/file_level_markers/function_level_markers. (5) still empty → check warnings/recommendations/patterns at top level. Canonical: {markers: [], total_markers: int, _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.8", "schema_normalizer.py", "Normalize file_dossiers.json",
     "file_dossiers.json", "file_dossiers.json (in-place)", "None ($0)",
     "8 key names: dossiers (dict or list), data_file_dossiers, file_dossiers, files, detailed_dossiers, code_file_dossiers, analysis_output_dossiers, conceptual_entity_dossiers, standard_dossiers, peripheral_files. Lists converted to dicts keyed by filename. Canonical: {dossiers: {filename: {summary, story_arc, key_decisions, warnings, related_files, confidence, claude_behavior, significance}}, total_files_cataloged: int, significance_tiers: {}, summary_statistics: {}, _extra: {}, _normalization_log: []}."),

    ("Schema Norm", "N.9", "schema_normalizer.py", "Normalize claude_md_analysis.json",
     "claude_md_analysis.json", "claude_md_analysis.json (in-place)", "None ($0)",
     "3 schema types: Schema A (file_analyses, key_findings_ranked, aggregate_statistics, session_overview), Schema B (gate_analysis, framing_analysis, claude_md_improvement_recommendations), Schema C (gate_activations, gates_not_triggered, overall_assessment, behavior_profile, tier2_security_analysis). Canonical: {gate_activations: [], gates_not_triggered: [], overall_assessment: {}, behavior_profile: {}, tier2_security_analysis: {}, _extra: {}, _normalization_log: []}."),

    ("LLM Pass 1", "L1", "llm_pass_runner.py", "Content-referential + assumption subtypes",
     "enriched_session.json (assistant messages)", "llm_pass_1_results.json", "Haiku 4.5",
     "PROMPT: 7-example system prompt distinguishing topic discussion from session dynamics. Per message outputs: {index, content_referential: bool, content_referential_reason: str, code_assumption: bool, format_assumption: bool, direction_assumption: bool, scope_assumption: bool, assumption_details: str}. Batched into prompt-sized chunks. Up to 50-60 concurrent Haiku calls. OPTIONAL — pipeline works without this."),

    ("LLM Pass 2", "L2", "llm_pass_runner.py", "Silent decisions + unverified claims + overconfidence",
     "enriched_session.json (assistant messages)", "llm_pass_2_results.json", "Opus 4.6",
     "Detects: silent_decision (Claude made a design choice without presenting options), unverified_claim (claimed something works without testing), overconfidence (expressed certainty beyond evidence). Up to 30 concurrent Opus calls. OPTIONAL."),

    ("LLM Pass 3", "L3", "llm_pass_runner.py", "Intent/assumption resolution (targeted)",
     "enriched_session.json + llm_pass_1_results.json", "llm_pass_3_results.json", "Opus 4.6",
     "Only processes messages flagged by Pass 1 with assumptions. Resolves ambiguous assumptions using surrounding user context. DEPENDS ON Pass 1 output. OPTIONAL."),

    ("LLM Pass 4", "L4", "llm_pass_runner.py", "Strategic importance scoring",
     "enriched_session.json", "llm_pass_4_results.json", "Haiku 4.5",
     "Assigns 1-10 importance score to each message based on strategic value for understanding the session's narrative arc. Independent of other passes. OPTIONAL."),

    ("LLM Merge", "L5", "merge_llm_results.py", "Merge all LLM pass results",
     "enriched_session.json + llm_pass_{1,2,3,4}_results.json", "enriched_session_v2.json", "None ($0)",
     "Adds llm_behavior field to each message record in enriched_session.json. Output is enriched_session_v2.json which prepare_agent_data.py prefers over v1."),

    ("Phase 1", "1.1", "extract_threads.py", "Load tier4 priority messages",
     "tier4_priority_messages.json", "", "None ($0)",
     "Reads {count: N, messages: [...]} from the Phase 0b output. These are the highest-signal messages only."),

    ("Phase 1", "1.2a", "extract_threads.py", "PER MESSAGE: Frustration detection",
     "", "", "None ($0)",
     "detect_frustration_level(msg) — scores 0-5 from: caps_ratio (>0.8=+3, >0.4=+2, >0.2=+1), profanity (+2), exclamations (>5=+2, >2=+1), emergency (+1), repeat_count (>3=+2, >1=+1). Content patterns: 'what the fuck' (+2), 'are you serious' (+2), specific profanity patterns (+2-3), 'rushing'/'ignored' (+1), 'dementia'/'moron' (+2), 'i really hate' (+1), 'like a god damn chump' (+2), 'all i care about' with caps (+2). Capped at 5."),

    ("Phase 1", "1.2b", "extract_threads.py", "PER MESSAGE: Deception detection",
     "", "", "None ($0)",
     "detect_deception(msg) — for assistant messages only. Checks: (1) overconfident=true AND damage_score>1, (2) 'comprehensive' + 'all' + 'complete' in content, (3) regex: 100%|everything.*fixed|all.*resolved, (4) failure_count>5 AND completion claims. Returns true if ≥2 patterns match."),

    ("Phase 1", "1.2c", "extract_threads.py", "PER MESSAGE: User ideas thread extraction",
     "", "", "None ($0)",
     "extract_user_ideas_thread(msg) — for user messages: detects 'I want/need' goals (evolution='expressing_goal'), 'what if' explorations ('exploring_new_direction'), corrections with 'instead/actually/no,' ('correcting_claude'), frustration keywords ('frustration_feedback'), continuation ('context_recovery'), architecture keywords ('architectural_thinking'), prevention keywords ('prevention_system_thinking'). For assistant messages: detects 'Your goal/vision/intent' interpretations. Reference session 3b7084d5 gets 14 hand-annotated ideas with specific evolution labels."),

    ("Phase 1", "1.2d", "extract_threads.py", "PER MESSAGE: Claude response thread extraction",
     "", "", "None ($0)",
     "extract_claude_response_thread(msg) — assistant only. Action: apologized_and_corrected, implemented, fixed, analyzed, executing_pipeline, reported_error, responded. Quality: harmful (damage≥3), poor (damage≥2, overconfident+failures, rushing, ignores_context, completion claims with failures>5), adequate (frustration>2 or failures>2), good. Pitch: offering_options, proposing_solution."),

    ("Phase 1", "1.2e", "extract_threads.py", "PER MESSAGE: Reactions thread extraction",
     "", "", "None ($0)",
     "extract_reactions_thread(msg) — user reactions: rage (caps>0.8 or profanity+caps>0.3), angry (profanity), emphatic (caps>0.3), frustrated (rushing/ignored/didn't listen), positive (great/perfect/excellent), correcting (instead/no/actually), session_continuation, questioning (?), neutral. Triggers: Sonnet substitution, rushing, truncation, deletion, summarization, fallbacks, metaphors, context loss ('dementia')."),

    ("Phase 1", "1.2f", "extract_threads.py", "PER MESSAGE: Software thread extraction",
     "", "", "None ($0)",
     "extract_software_thread(msg) — from metadata: files_create (created), files_edit (edited), files (mentioned). From content: deletion keywords + file name matching. Output: {created: [], modified: [] (capped at 10), deleted: []}."),

    ("Phase 1", "1.2g", "extract_threads.py", "PER MESSAGE: Code blocks thread extraction",
     "", "", "None ($0)",
     "extract_code_blocks_thread(msg) — regex for code blocks (```...```), function names (`func()`), class names (patterns: *Extractor, *Builder, *Analyzer, etc.). Action from content: reverted/fixed/created/modified/analyzed/removed/referenced. Output: {action: str, blocks: [] (capped at 8), code_block_count: int}."),

    ("Phase 1", "1.2h", "extract_threads.py", "PER MESSAGE: Plans thread extraction",
     "", "", "None ($0)",
     "extract_plans_thread(msg) — from filter_signals: plan score sum. If >0: regex for checkmarks (✅✓ = completed), x-marks (❌⬜☐ = pending), task references (Task #N), phase references (Phase N). Output: {detected: bool, content: str, completed: [] (capped at 5), pending: [] (capped at 5)}."),

    ("Phase 1", "1.2i", "extract_threads.py", "PER MESSAGE: 6 marker extraction",
     "", "", "None ($0)",
     "extract_markers(msg) — is_pivot (pivot_score≥3 or keywords: instead, new approach, different strategy, pivot, change of plan, scrap that), is_failure (failure_score≥5 or: devastating, broken, crashed, failed, api error — assistant only), is_breakthrough (breakthrough_score≥2 or: it works!, working!, eureka, the fix is, smoking gun, root cause), is_ignored_gem (user messages with architectural/insight patterns — 36 gem keyword patterns), deception_detected (from 1.2b), frustration_level (from 1.2a)."),

    ("Phase 1", "1.3", "extract_threads.py", "Post-process: verify ignored gems",
     "", "", "None ($0)",
     "For each user message marked is_ignored_gem=true: check if the NEXT Claude message addresses the key terms (extract top 5 words >5 chars from user message, check ≥50% appear in Claude's response). If addressed: unmark is_ignored_gem=false."),

    ("Phase 1", "1.4", "extract_threads.py", "Build canonical thread output",
     "", "", "None ($0)",
     "Groups all per-message extractions into 6 canonical thread categories: ideas (from user_ideas where idea is not null), reactions (where type is not neutral/null), software (where files exist), code (where blocks or code_block_count>0), plans (where detected=true), behavior (where quality is harmful/poor, or any marker is true). Each entry: {msg_index, content (capped at 500), significance (high/medium)}. Full per-message extractions preserved in _extra.extractions."),

    ("Phase 1", "1.5", "extract_threads.py", "Write thread_extractions.json",
     "", "thread_extractions.json", "None ($0)",
     "Output: {session_id: str, threads: {ideas: {description, entries}, reactions: {description, entries}, software: {description, entries}, code: {description, entries}, plans: {description, entries}, behavior: {description, entries}}, _extra: {total_analyzed: int, extraction_method: 'deterministic_pattern_matching', extractions: [full per-message data]}}. Typically 50-500 KB."),

    ("Phase 1+", "1+.1", "batch_orchestrator.py", "Opus agent: Thread Analyst",
     "safe_condensed.json", "thread_extractions.json (LLM version)", "Opus 4.6",
     "LLM-based thread extraction. Richer semantic analysis than the deterministic version — can understand context, sarcasm, implicit meaning. Produces same canonical format. OPTIONAL — only runs in --full mode."),

    ("Phase 1+", "1+.2", "batch_orchestrator.py", "Opus agent: Primitives Tagger",
     "safe_condensed.json", "semantic_primitives.json", "Opus 4.6",
     "Tags each message with 7 semantic primitives: action_vector (building/investigating/defending/recovering/...), confidence_signal (working/stable/fragile/tentative/proven), emotional_tenor (frustrated/curious/relieved/focused/...), intent_marker (exploring/correcting/demanding/teaching/...), friction_log (text description of friction), decision_trace (what was decided and why), disclosure_pointer (what was revealed about the human-AI dynamic). Plus session-level distributions and summary. OPTIONAL."),

    ("Phase 1+", "1+.3", "batch_orchestrator.py", "Opus agent: Geological Reader",
     "safe_condensed.json", "geological_notes.json", "Opus 4.6",
     "Multi-resolution reading at 3 zoom levels: micro (individual messages — specific moments, line-level observations), meso (5-message windows — patterns forming, interactions between steps), macro (15-20 message arcs — narrative structures, emotional trajectories, phase transitions). Also: geological_metaphor (session as geological formation), session_character. OPTIONAL."),

    ("Phase 1+", "1+.4", "batch_orchestrator.py", "Opus agent: Free Explorer",
     "safe_condensed.json", "explorer_notes.json", "Opus 4.6",
     "Unconstrained observation agent — no structured output format required. Can note: patterns, anomalies, abandoned ideas, cross-session connections, unanswered questions, emotional dynamics, what matters most, data quality issues. The explorer's value is in noticing what structured extractors miss. OPTIONAL."),

    ("Phase 2", "2.1", "batch_p2_generator.py", "Build idea graph",
     "session_metadata.json + thread_extractions.json + geological_notes.json + semantic_primitives.json + explorer_notes.json", "idea_graph.json", "None ($0)",
     "Extracts idea nodes from thread entries (canonical format: categories with entries). Each node: {id, name (first 80 chars of content), description (full content), first_appearance (msg_index), confidence, emotional_context, trigger}. Edges between sequential nodes: {from_id, to_id, transition_type='evolved', trigger_message, evidence}. Handles both canonical and old extraction formats. Output: {nodes: [], edges: [], metadata: {}}. Typically 20-200 KB."),

    ("Phase 2", "2.2", "batch_p2_generator.py", "Build synthesis",
     "Same 5 inputs + idea_graph.json", "synthesis.json", "None ($0)",
     "Multi-pass synthesis — conceptually 6 passes at increasing 'temperatures': factual (0.3), patterns (0.5), vertical structures (0.7), creative synthesis (0.9), wild connections (1.0), grounding (0.0). Each pass builds on prior passes. Session character summary. Output: {session_id, passes: {}, key_findings: [], session_character: str, cross_session_links: []}. Typically 10-100 KB."),

    ("Phase 2", "2.3", "batch_p2_generator.py", "Build grounded markers",
     "Same 5 inputs + idea_graph.json + synthesis.json", "grounded_markers.json", "None ($0)",
     "Generates verifiable markers: warnings (W01-W12), recommendations (R01-R12), behavioral patterns (B01-B08), iron rules. Each marker: {marker_id, category, claim (verifiable statement), confidence (0-1), evidence (specific reference), source (session:msg_index), severity}. Output: {markers: [], total_markers: int}. Typically 10-100 KB."),

    ("Phase 3a", "3a.1", "collect_file_evidence.py", "Load 11 data sources",
     "session_metadata.json + geological_notes.json + semantic_primitives.json + explorer_notes.json + file_genealogy.json + thread_extractions.json + idea_graph.json + grounded_markers.json + synthesis.json + claude_md_analysis.json + file_dossiers.json",
     "", "None ($0)",
     "load_json(filename, search_dirs) — searches PERMANENT_HYPERDOCS/sessions/session_X first (rich Opus data), then local output/session_X. Returns {} if not found. Loads all 11 data sources into memory."),

    ("Phase 3a", "3a.2", "collect_file_evidence.py", "Get target files from session",
     "", "", "None ($0)",
     "get_all_target_files(session_metadata) — reads top_files (list of [filename, count] pairs) and file_mention_counts (dict). Returns sorted unique filenames. Typically 10-40 files per session. Some sessions have 200+ files (e.g., session 2146922a had 822 evidence files)."),

    ("Phase 3a", "3a.3", "collect_file_evidence.py", "PER FILE: Find mention indices",
     "", "", "None ($0)",
     "find_mention_indices() — scans 3 data sources: (1) thread_extractions: handles both dict and list schemas, checks content field of entries for filename mentions. (2) file_dossiers: reads first_mention_index, last_mention_index, mentioned_in array. Handles both dict and list dossier schemas. (3) geological_notes: checks all zoom levels for filename mentions, expands message_range arrays into index sets."),

    ("Phase 3a", "3a.4", "collect_file_evidence.py", "PER FILE: Build time window",
     "", "", "None ($0)",
     "build_window_indices(mention_indices, window=10) — expands each mention index by ±10 messages. Returns set of all non-negative indices in the window. This is the 'context window' that Phase 3a uses to correlate emotional/behavioral data with file references."),

    ("Phase 3a", "3a.5", "collect_file_evidence.py", "PER FILE: Build 11 evidence sections",
     "", "", "None ($0)",
     "11 sections built per file:\n(1) emotional_arc: semantic primitives in the ±10 window — emotional_tenor, confidence_signal, action_vector, intent_marker, friction_log, decision_trace per message. Plus 4 distributions, file-window distribution, emotion trajectory.\n(2) geological_character: micro/meso/macro observations matching by filename or time-window overlap. Match reason tagged: 'filename' or 'time_window'.\n(3) lineage: file genealogy family, idea graph nodes mentioning the file, dossier extras (alternate_locations, key_insight, role_in_session, etc.).\n(4) explorer_observations: matching observations, verification issues, anomalies, data quality.\n(5) chronological_timeline: thread entries + grounded markers mentioning the file, sorted by msg_index, deduplicated by (msg_index, thread).\n(6) code_similarity: placeholder — populated by Phase 4a.\n(7) graph_context: nodes mentioning file (with full details), connected edges, containing subgraphs, session graph statistics.\n(8) synthesis_context: all 6 synthesis passes with sub-keys, cross_pass_summary, key_findings, cross_session_links.\n(9) claude_md_context: gate activations, findings, recommendations, behavior profile, tier2 security — handles 3 schema variants.\n(10) session_context: message counts, tokens, tier distribution, frustration peaks (session-level).\n(11) thread_markers: pivot/failure/breakthrough/ignored_gem/deception/frustration markers from messages mentioning this file."),

    ("Phase 3a", "3a.6", "collect_file_evidence.py", "PER FILE: Write evidence JSON",
     "", "file_evidence/{safe_filename}_evidence.json", "None ($0)",
     "One JSON per target file. safe_filename converts / \\ . space to _. Format: {file: str, session: str, mention_indices: [], window_size: 10, emotional_arc: {...}, geological_character: {...}, lineage: {...}, explorer_observations: {...}, chronological_timeline: {...}, code_similarity: {matches: [], data_points: 0}, graph_context: {...}, synthesis_context: {...}, claude_md_context: {...}, session_context: {...}, thread_markers: {...}}. Typically 5-50 KB per file."),

    ("Phase 3b", "3b.1", "generate_dossiers.py", "Load session data",
     "session_metadata.json + grounded_markers.json + thread_extractions.json + idea_graph.json", "", "None ($0)",
     "Loads 4 data sources for dossier generation."),

    ("Phase 3b", "3b.2", "generate_dossiers.py", "Build file dossiers (top 15 files)",
     "", "file_dossiers.json", "None ($0)",
     "Top 15 files by mention count. Per file: {file_name, total_mentions, first_mention_index, last_mention_index, summary, story_arc, key_decisions: [], warnings: [], related_files: [], confidence, claude_behavior: {}, significance (tier), grounded_markers: [], idea_graph_nodes: [], source_messages: []}. Format: {dossiers: {filename: {...}}, total_files_cataloged: int, significance_tiers: {}}. Typically 20-200 KB."),

    ("Phase 3b", "3b.3", "generate_dossiers.py", "Build claude_md_analysis",
     "", "claude_md_analysis.json", "None ($0)",
     "How CLAUDE.md gates affected this session. Gate activations, behavior profile, security analysis. Typically 5-50 KB."),

    ("Phase 3c", "3c.1", "generate_viewer.py", "Load all 10 pipeline outputs",
     "session_metadata + thread_extractions + geological_notes + semantic_primitives + explorer_notes + idea_graph + synthesis + grounded_markers + file_dossiers + claude_md_analysis", "", "None ($0)",
     "Loads everything for a single-page HTML visualization."),

    ("Phase 3c", "3c.2", "generate_viewer.py", "Generate session HTML viewer",
     "", "session_viewer.html", "None ($0)",
     "Single HTML page with tabs: session stats (message counts, tiers, frustration peaks), threads (6 categories), geological (micro/meso/macro), primitives (distributions, per-message), explorer (observations, verification), idea graph (nodes, edges, subgraphs), synthesis (6 passes), markers (warnings, recommendations), dossiers (per-file), claude_md analysis. Typically 100-500 KB."),

    ("Phase 4a", "4a.1", "aggregate_dossiers.py", "Read ALL session dossiers (277 sessions)",
     "output/session_*/file_dossiers.json (ALL sessions)", "", "None ($0)",
     "Iterates every session_* directory. For each: loads file_dossiers.json, detects dict vs list schema, normalizes into {file, session_id, confidence, total_mentions, story_arc, key_decisions, warnings, related_files, claude_behavior, significance, grounded_markers, idea_graph_nodes, source_messages}. Counts: 277 sessions read, 161 dict format, 116 list format, 21 skipped, 4464 total dossier entries."),

    ("Phase 4a", "4a.2", "aggregate_dossiers.py", "Group by file path → aggregate",
     "", "", "None ($0)",
     "Groups 4464 entries by filename → 1379 unique files. Per file: {file_path, session_count, total_mentions, exists_on_disk, sessions: [], merged_warnings: [], merged_decisions: [], confidence_history: [{session, confidence}], story_arcs: [{session, arc}], behavioral_patterns: {pattern: [{session, value}]}, significance_scores: {tier: count}, related_files: []}. 690 single-session files, 233 two-session, 456 three+ sessions. 478 exist on disk."),

    ("Phase 4a", "4a.3", "aggregate_dossiers.py", "Load code similarity index",
     "code_similarity_index.json (8.1 MB, 14176 matches)", "", "None ($0)",
     "Builds per-file lookup. For each match: extracts file_a, file_b, score (from signals.signal_score or text_similarity or func_overlap), pattern_type. Filters score>0.3, keeps top 10 per file sorted by score desc. Builds basename lookup so both full paths and bare filenames resolve. 404 files get matches."),

    ("Phase 4a", "4a.4", "aggregate_dossiers.py", "Load genealogy families",
     "file_genealogy.json (from all session dirs)", "", "None ($0)",
     "Scans all session_*/file_genealogy.json. Per family: concept name, file members, session source. 15 files in families from 15 mapped."),

    ("Phase 4a", "4a.5", "aggregate_dossiers.py", "Build frustration-file map",
     "session_metadata.json + enriched_session.json (per session)", "", "None ($0)",
     "For each session with frustration_peaks: loads enriched_session to get per-message file mentions. For each peak: finds files mentioned within ±5 messages. Output per file: [{session, message_index, caps_ratio, profanity}]. 623 files associated with frustration peaks."),

    ("Phase 4a", "4a.6", "aggregate_dossiers.py", "Load cross-session evidence (53 sessions)",
     "session_*/file_evidence/*_evidence.json (ALL sessions)", "", "None ($0)",
     "Scans ALL session directories for file_evidence/ subdirs. For each evidence file: aggregates 8 cross-session data types per filename. (1) cross_session_emotional_arc: per_session_emotions + cross_emotion_distribution. (2) cross_session_geological: all observations tagged by session and zoom. (3) cross_session_explorer: all observations tagged by session. (4) cross_session_timeline: all events tagged by session. (5) cross_session_graph_context: per_session graph stats + all_connected_edges + all_subgraphs. (6) cross_session_synthesis: per_session passes and findings. (7) cross_session_geological_metaphors: session→metaphor map. (8) cross_session_claude_md: per_session analysis. From 53 sessions, 2641 evidence files → 812 files enriched."),

    ("Phase 4a", "4a.7", "aggregate_dossiers.py", "Write cross_session_file_index.json",
     "", "cross_session_file_index.json (98 MB)", "None ($0)",
     "Per file (1379 entries): file_path, session_count, total_mentions, exists_on_disk, sessions, merged_warnings, merged_decisions, confidence_history, story_arcs, behavioral_patterns, significance_scores, related_files, code_similarity, genealogy, frustration_associations, cross_session_emotional_arc, cross_session_geological, cross_session_explorer, cross_session_timeline, cross_session_graph_context, cross_session_synthesis, cross_session_geological_metaphors, cross_session_claude_md. Also writes to PERMANENT_HYPERDOCS/indexes/."),

    ("Phase 4a", "4a.8", "aggregate_dossiers.py", "Write per-file extracts to hyperdoc_inputs/",
     "", "hyperdoc_inputs/{safe_filename}.json (456 files)", "None ($0)",
     "One JSON per file with ≥3 sessions. Same content as the cross_session_file_index entry for that file. Written to both output/hyperdoc_inputs/ and PERMANENT_HYPERDOCS/hyperdoc_inputs/."),
]


def build_sheet_1_pipeline_steps(wb):
    """Sheet 1: Every single pipeline step."""
    ws = wb.create_sheet("Pipeline Steps")
    headers = ["Phase", "Step ID", "Script", "Operation", "Reads", "Writes", "API / Cost", "Exhaustive Detail"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    prev_phase = ""
    for phase, step_id, script, operation, reads, writes, api, detail in PIPELINE_STEPS:
        if phase != prev_phase:
            # Phase separator row
            ws.cell(row=row, column=1, value=phase)
            style_phase_row(ws, row, len(headers))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            row += 1
            prev_phase = phase

        ws.cell(row=row, column=1, value=phase)
        ws.cell(row=row, column=2, value=step_id)
        ws.cell(row=row, column=3, value=script)
        ws.cell(row=row, column=4, value=operation)
        ws.cell(row=row, column=5, value=reads)
        ws.cell(row=row, column=6, value=writes)
        ws.cell(row=row, column=7, value=api)
        ws.cell(row=row, column=8, value=detail)

        # Color the API column
        api_cell = ws.cell(row=row, column=7)
        if "None" in api or "$0" in api:
            api_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        elif "Opus" in api:
            api_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        elif "Haiku" in api:
            api_cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = WRAP
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Set column widths
    widths = [14, 8, 28, 40, 45, 40, 14, 100]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    return ws


def build_sheet_2_file_schemas(wb):
    """Sheet 2: Every field in every JSON file type, with types and samples from a real session."""
    ws = wb.create_sheet("File Schemas (Field-Level)")
    headers = ["File", "JSON Path", "Type", "Depth", "Sample Value / Description"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # Pick a data-rich session to sample
    sample_session = PERM_SESSIONS / "session_0012ebed"
    if not sample_session.exists():
        # Fallback to any session with rich data
        for d in sorted(PERM_SESSIONS.iterdir()):
            if d.is_dir() and (d / "session_metadata.json").exists() and (d / "idea_graph.json").exists():
                sample_session = d
                break

    files_to_scan = [
        "session_metadata.json", "enriched_session.json",
        "thread_extractions.json", "semantic_primitives.json",
        "geological_notes.json", "explorer_notes.json",
        "idea_graph.json", "synthesis.json", "grounded_markers.json",
        "file_dossiers.json", "claude_md_analysis.json",
    ]

    # Also check for evidence files
    ev_dir = OUTPUT_DIR / sample_session.name / "file_evidence"
    if not ev_dir.exists():
        ev_dir = sample_session / "file_evidence"
    sample_evidence = None
    if ev_dir.exists():
        ev_files = sorted(ev_dir.glob("*_evidence.json"))
        if ev_files:
            files_to_scan.append(f"file_evidence/{ev_files[0].name}")
            sample_evidence = ev_files[0]

    row = 2
    for filename in files_to_scan:
        # Phase separator
        ws.cell(row=row, column=1, value=filename)
        style_phase_row(ws, row, len(headers))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1

        # Load the file
        if "file_evidence" in filename and sample_evidence:
            data = load_json_safe(sample_evidence)
        else:
            # Check both local output and PERM
            data = load_json_safe(OUTPUT_DIR / sample_session.name / filename)
            if data is None:
                data = load_json_safe(sample_session / filename)

        if data is None:
            ws.cell(row=row, column=2, value="(file not found or invalid JSON)")
            row += 1
            continue

        # For enriched_session.json, only show first message to avoid explosion
        if filename == "enriched_session.json" and "messages" in data and isinstance(data["messages"], list):
            # Show top-level keys first
            top_fields = flatten_json_fields({k: v for k, v in data.items() if k != "messages"}, max_depth=2)
            for path, ftype, sample, depth in top_fields:
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=path)
                ws.cell(row=row, column=3, value=ftype)
                ws.cell(row=row, column=4, value=depth)
                ws.cell(row=row, column=5, value=sample[:300])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).alignment = WRAP
                    ws.cell(row=row, column=col).border = THIN_BORDER
                    ftype_color = FIELD_TYPE_COLORS.get(ftype, "ffffff")
                    ws.cell(row=row, column=3).fill = PatternFill(start_color=ftype_color, end_color=ftype_color, fill_type="solid")
                row += 1

            # Then show first message structure
            if data["messages"]:
                ws.cell(row=row, column=2, value="messages[0] (structure of one enriched message)")
                style_phase_row(ws, row, len(headers))
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
                row += 1
                msg_fields = flatten_json_fields(data["messages"][0], "messages[0]", max_depth=3)
                for path, ftype, sample, depth in msg_fields:
                    ws.cell(row=row, column=1, value=filename)
                    ws.cell(row=row, column=2, value=path)
                    ws.cell(row=row, column=3, value=ftype)
                    ws.cell(row=row, column=4, value=depth)
                    ws.cell(row=row, column=5, value=sample[:300])
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).alignment = WRAP
                        ws.cell(row=row, column=col).border = THIN_BORDER
                    row += 1
        else:
            fields = flatten_json_fields(data, max_depth=3)
            for path, ftype, sample, depth in fields:
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=path)
                ws.cell(row=row, column=3, value=ftype)
                ws.cell(row=row, column=4, value=depth)
                ws.cell(row=row, column=5, value=sample[:300])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).alignment = WRAP
                    ws.cell(row=row, column=col).border = THIN_BORDER
                    ftype_color = FIELD_TYPE_COLORS.get(ftype, "ffffff")
                    ws.cell(row=row, column=3).fill = PatternFill(start_color=ftype_color, end_color=ftype_color, fill_type="solid")
                row += 1

    widths = [30, 50, 8, 6, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def build_sheet_3_session_inventory(wb):
    """Sheet 3: Per-session file inventory for all processed sessions."""
    ws = wb.create_sheet("Session Inventory (All 292)")
    headers = [
        "Session ID", "Location",
        "enriched_session", "session_metadata", "tier4_priority",
        "safe_tier4", "safe_condensed", "batches",
        "thread_extractions", "semantic_primitives",
        "geological_notes", "explorer_notes",
        "idea_graph", "synthesis", "grounded_markers",
        "file_evidence/", "evidence_file_count",
        "file_dossiers", "claude_md_analysis",
        "session_viewer",
        "Total Files", "Total Size (KB)"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    seen = set()
    for base in [PERM_SESSIONS, OUTPUT_DIR]:
        if not base.exists():
            continue
        for d in sorted(base.iterdir()):
            if not d.is_dir() or not d.name.startswith("session_"):
                continue
            sid = d.name.replace("session_", "")[:8]
            if sid in seen:
                continue
            seen.add(sid)

            loc = "PERM" if "PERMANENT" in str(base) else "LOCAL"

            # Check each file
            file_checks = [
                "enriched_session.json", "session_metadata.json",
                "tier4_priority_messages.json", "safe_tier4.json",
                "safe_condensed.json", "batches",
                "thread_extractions.json", "semantic_primitives.json",
                "geological_notes.json", "explorer_notes.json",
                "idea_graph.json", "synthesis.json", "grounded_markers.json",
                "file_evidence", None,  # placeholder for evidence count
                "file_dossiers.json", "claude_md_analysis.json",
            ]

            ws.cell(row=row, column=1, value=sid)
            ws.cell(row=row, column=2, value=loc)

            total_files = 0
            total_size = 0
            col_idx = 3
            for check in file_checks:
                if check is None:
                    # Evidence count
                    ev_dir = d / "file_evidence"
                    # Also check local
                    ev_dir_local = OUTPUT_DIR / f"session_{sid}" / "file_evidence"
                    ev_count = 0
                    for ed in [ev_dir, ev_dir_local]:
                        if ed.exists():
                            ev_count = len(list(ed.glob("*_evidence.json")))
                            break
                    ws.cell(row=row, column=col_idx, value=ev_count)
                    col_idx += 1
                    continue

                path = d / check
                # Also check local output
                path_local = OUTPUT_DIR / f"session_{sid}" / check
                exists = path.exists() or path_local.exists()

                if exists:
                    ws.cell(row=row, column=col_idx, value="Y")
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color="d4edda", end_color="d4edda", fill_type="solid")
                    total_files += 1
                    actual_path = path if path.exists() else path_local
                    if actual_path.is_file():
                        total_size += actual_path.stat().st_size // 1024
                else:
                    ws.cell(row=row, column=col_idx, value="-")
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                col_idx += 1

            # Check viewer
            viewer_exists = (d / "session_viewer.html").exists() or len(list(d.glob("*viewer*.html"))) > 0
            ws.cell(row=row, column=col_idx, value="Y" if viewer_exists else "-")
            if viewer_exists:
                ws.cell(row=row, column=col_idx).fill = PatternFill(
                    start_color="d4edda", end_color="d4edda", fill_type="solid")
            else:
                ws.cell(row=row, column=col_idx).fill = PatternFill(
                    start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            col_idx += 1

            ws.cell(row=row, column=col_idx, value=total_files)
            col_idx += 1
            ws.cell(row=row, column=col_idx, value=total_size)

            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")

            row += 1

    ws.freeze_panes = "C2"
    auto_width(ws, min_width=8, max_width=20)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    return ws


def build_sheet_4_cross_session_top_files(wb):
    """Sheet 4: Top files from cross-session index with all their data."""
    ws = wb.create_sheet("Cross-Session Top Files")
    headers = [
        "File", "Sessions", "Mentions", "On Disk",
        "Warnings", "Decisions", "Confidence History",
        "Code Sim Matches", "Genealogy Family",
        "Frustration Assocs", "Story Arcs",
        "Emo Sessions", "Geo Observations",
        "Explorer Notes", "Timeline Events",
        "Graph Edges", "Synthesis Sessions",
        "Geo Metaphors", "Claude MD Sessions"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # Load the index
    idx_path = OUTPUT_DIR / "cross_session_file_index.json"
    if not idx_path.exists():
        idx_path = INDEXES_DIR / "cross_session_file_index.json"
    if not idx_path.exists():
        ws.cell(row=2, column=1, value="cross_session_file_index.json not found")
        return ws

    data = load_json_safe(idx_path)
    if not data:
        ws.cell(row=2, column=1, value="Failed to load index")
        return ws

    files = data.get("files", {})
    # Sort by session_count desc
    sorted_files = sorted(files.items(), key=lambda x: x[1].get("session_count", 0), reverse=True)

    row = 2
    for filepath, entry in sorted_files[:200]:  # Top 200
        emo_arc = entry.get("cross_session_emotional_arc", {})
        geo = entry.get("cross_session_geological", [])
        explorer = entry.get("cross_session_explorer", [])
        timeline = entry.get("cross_session_timeline", [])
        gc = entry.get("cross_session_graph_context", {})
        synth = entry.get("cross_session_synthesis", {})
        metaphors = entry.get("cross_session_geological_metaphors", {})
        cmd = entry.get("cross_session_claude_md", {})

        gen = entry.get("genealogy")
        gen_name = gen.get("family_name", "") if gen else ""

        conf_hist = entry.get("confidence_history", [])
        conf_str = " → ".join([str(c.get("confidence", "?")) for c in conf_hist[:8]])
        if len(conf_hist) > 8:
            conf_str += "..."

        vals = [
            filepath,
            entry.get("session_count", 0),
            entry.get("total_mentions", 0),
            "Y" if entry.get("exists_on_disk") else "",
            len(entry.get("merged_warnings", [])),
            len(entry.get("merged_decisions", [])),
            conf_str,
            len(entry.get("code_similarity", [])),
            gen_name,
            len(entry.get("frustration_associations", [])),
            len(entry.get("story_arcs", [])),
            len(emo_arc.get("per_session_emotions", {})),
            len(geo),
            len(explorer),
            len(timeline),
            len(gc.get("all_connected_edges", [])),
            len(synth.get("per_session", {})),
            len(metaphors),
            len(cmd.get("per_session", {})),
        ]

        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val)
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = WRAP

        row += 1

    ws.freeze_panes = "B2"
    auto_width(ws, min_width=8, max_width=30)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["G"].width = 50
    return ws


def build_sheet_5_data_flow(wb):
    """Sheet 5: Which file feeds which file — the complete dependency graph."""
    ws = wb.create_sheet("Data Flow (Who Reads What)")
    headers = ["Consumer (Script)", "Reads This File", "Produces This File", "Relationship"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    flows = [
        ("deterministic_prep.py", "{session_id}.jsonl", "enriched_session.json", "Raw JSONL → enriched messages with 18 fields per message"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "session_metadata.json", "Strip messages array, keep stats only"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "tier2plus_messages.json", "Filter: keep messages where filter_tier >= 2"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "tier4_priority_messages.json", "Filter: keep messages where filter_tier == 4"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "conversation_condensed.json", "All messages, compressed key names"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "user_messages_tier2plus.json", "Filter: user role AND tier >= 2"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "emergency_contexts.json", "Window of ±5 msgs around each emergency"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "batches/batch_*.json", "Split tier 2+ into chunks of 30"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "safe_tier4.json", "Profanity-sanitized tier 4 messages"),
        ("prepare_agent_data.py", "enriched_session.json (or v2)", "safe_condensed.json", "Profanity-sanitized all messages"),
        ("schema_normalizer.py", "thread_extractions.json", "thread_extractions.json (normalized)", "Fix 5 schema variants → canonical {threads: {cat: {entries}}}"),
        ("schema_normalizer.py", "semantic_primitives.json", "semantic_primitives.json (normalized)", "Fix 5 schema variants → canonical {tagged_messages: []}"),
        ("schema_normalizer.py", "geological_notes.json", "geological_notes.json (normalized)", "Fix 14 key variants → canonical {micro, meso, macro, observations}"),
        ("schema_normalizer.py", "explorer_notes.json", "explorer_notes.json (normalized)", "Fix 12 key variants → canonical {observations, explorer_summary}"),
        ("schema_normalizer.py", "idea_graph.json", "idea_graph.json (normalized)", "Fix 3 schema variants → canonical {nodes, edges, metadata}"),
        ("schema_normalizer.py", "synthesis.json", "synthesis.json (normalized)", "Fix 5 schema variants → canonical {passes, key_findings}"),
        ("schema_normalizer.py", "grounded_markers.json", "grounded_markers.json (normalized)", "Fix 6 schema variants → canonical {markers, total_markers}"),
        ("schema_normalizer.py", "file_dossiers.json", "file_dossiers.json (normalized)", "Fix 8 key variants → canonical {dossiers: {file: {...}}}"),
        ("schema_normalizer.py", "claude_md_analysis.json", "claude_md_analysis.json (normalized)", "Fix 3 schema types → canonical {gate_activations, overall_assessment}"),
        ("llm_pass_runner.py", "enriched_session.json", "llm_pass_1_results.json", "Haiku: content-referential + assumptions"),
        ("llm_pass_runner.py", "enriched_session.json", "llm_pass_2_results.json", "Opus: silent decisions + unverified claims"),
        ("llm_pass_runner.py", "enriched_session.json + pass 1 results", "llm_pass_3_results.json", "Opus: intent/assumption resolution"),
        ("llm_pass_runner.py", "enriched_session.json", "llm_pass_4_results.json", "Haiku: strategic importance scoring"),
        ("merge_llm_results.py", "enriched_session.json + all pass results", "enriched_session_v2.json", "Merge LLM behavioral tags into enriched messages"),
        ("extract_threads.py", "tier4_priority_messages.json", "thread_extractions.json", "6 threads + 6 markers per tier-4 message"),
        ("Opus agent (Thread Analyst)", "safe_condensed.json", "thread_extractions.json", "LLM thread extraction (optional, richer)"),
        ("Opus agent (Primitives Tagger)", "safe_condensed.json", "semantic_primitives.json", "7 semantic primitives per message"),
        ("Opus agent (Geological Reader)", "safe_condensed.json", "geological_notes.json", "Multi-resolution observations"),
        ("Opus agent (Free Explorer)", "safe_condensed.json", "explorer_notes.json", "Free-form observations and anomalies"),
        ("batch_p2_generator.py", "session_metadata + thread_extractions + geological_notes + semantic_primitives + explorer_notes", "idea_graph.json", "Ideas as nodes, transitions as edges"),
        ("batch_p2_generator.py", "Same 5 inputs + idea_graph", "synthesis.json", "6-pass multi-temperature synthesis"),
        ("batch_p2_generator.py", "Same 5 inputs + idea_graph + synthesis", "grounded_markers.json", "Verifiable markers: warnings, recommendations"),
        ("collect_file_evidence.py", "11 data sources (metadata, threads, geo, prims, explorer, genealogy, idea_graph, markers, synthesis, claude_md, dossiers)", "file_evidence/*_evidence.json", "11 evidence sections per file, ±10 msg time window"),
        ("generate_dossiers.py", "session_metadata + markers + threads + idea_graph", "file_dossiers.json", "Top 15 files: full behavioral/structural profiles"),
        ("generate_dossiers.py", "Same inputs", "claude_md_analysis.json", "How CLAUDE.md gates affected this session"),
        ("generate_viewer.py", "All 10 pipeline JSON files", "session_viewer.html", "Single-page HTML visualization"),
        ("aggregate_dossiers.py", "ALL session_*/file_dossiers.json", "cross_session_file_index.json", "Per-file aggregation across 277 sessions"),
        ("aggregate_dossiers.py", "code_similarity_index.json", "cross_session_file_index.json (code_similarity field)", "14,176 AST matches → per-file top 10"),
        ("aggregate_dossiers.py", "ALL session_*/file_genealogy.json", "cross_session_file_index.json (genealogy field)", "18 file families"),
        ("aggregate_dossiers.py", "ALL session_*/session_metadata.json + enriched_session.json", "cross_session_file_index.json (frustration_associations field)", "Files ±5 msgs from frustration peaks"),
        ("aggregate_dossiers.py", "ALL session_*/file_evidence/*_evidence.json", "cross_session_file_index.json (8 cross_session_* fields)", "53 sessions × 2641 evidence files → 812 files enriched"),
        ("aggregate_dossiers.py", "cross_session_file_index.json", "hyperdoc_inputs/*.json", "Per-file extracts for files with 3+ sessions"),
        ("NOTHING", "cross_session_file_index.json", "⚠ NOWHERE — THE GAP", "Phase 4a output is never consumed. Phase 1 agents analyze blind."),
    ]

    row = 2
    for consumer, reads, writes, rel in flows:
        ws.cell(row=row, column=1, value=consumer)
        ws.cell(row=row, column=2, value=reads)
        ws.cell(row=row, column=3, value=writes)
        ws.cell(row=row, column=4, value=rel)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = WRAP
            ws.cell(row=row, column=col).border = THIN_BORDER

        # Highlight the gap
        if "GAP" in writes:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = GAP_FILL
                ws.cell(row=row, column=col).font = GAP_FONT
        row += 1

    widths = [35, 55, 50, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def main():
    print("=" * 60)
    print("Generating Exhaustive Pipeline Excel")
    print("=" * 60)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Sheet 1: Pipeline Steps (every micro-operation)...")
    build_sheet_1_pipeline_steps(wb)
    print(f"  {len(PIPELINE_STEPS)} steps documented")

    print("Sheet 2: File Schemas (field-level detail from real data)...")
    build_sheet_2_file_schemas(wb)

    print("Sheet 3: Session Inventory (all 292 sessions)...")
    build_sheet_3_session_inventory(wb)

    print("Sheet 4: Cross-Session Top Files (from index)...")
    build_sheet_4_cross_session_top_files(wb)

    print("Sheet 5: Data Flow (complete dependency graph)...")
    build_sheet_5_data_flow(wb)

    # Save
    out_path = H3 / "experiment" / "feedback_loop" / "output" / "pipeline_complete_anatomy.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    size = out_path.stat().st_size
    print(f"\nOutput: {out_path}")
    print(f"Size: {size:,} bytes")
    print(f"Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}: {ws.max_row} rows × {ws.max_column} cols")


if __name__ == "__main__":
    main()
